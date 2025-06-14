# mbe_recompiler.py
import struct
import openpyxl
import os
import argparse
import re

# --- MUDANÇA AQUI: Adicionado novo tipo de coluna ---
TYPE_NAME_TO_CODE = {"Int": 0x2, "String": 0x7, "StringID": 0x8, "IntID": 0x9}

def get_padded_string(text):
    if text is None: text = ""
    encoded = str(text).encode('utf-8')
    padded = encoded + b'\x00\x00'
    while len(padded) % 4 != 0: padded += b'\x00'
    return padded, len(padded)

def recompile_mbe(xlsx_path, mbe_path):
    print(f"Iniciando recompilação de '{xlsx_path}'...")
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        header_pattern = re.compile(r'(\w+)\s\((\d+)\)')
        column_types = [TYPE_NAME_TO_CODE[header_pattern.match(str(c.value)).group(1)] for c in ws[1] if c.value]
        
        all_rows_data = []
        for row in ws.iter_rows(min_row=2):
            row_data = []
            for cell, col_type in zip(row, column_types):
                # O 'else' já lida com qualquer tipo numérico (Int, IntID)
                if col_type in [TYPE_NAME_TO_CODE["String"], TYPE_NAME_TO_CODE["StringID"]]:
                    row_data.append(str(cell.value) if cell.value is not None else "")
                else:
                    row_data.append(int(cell.value) if cell.value is not None else 0)
            all_rows_data.append(row_data)
        
        # A regra de padding 0xCC é específica para 'Int' (0x2), não afeta 'IntID' (0x9)
        use_cc_padding_after_first_int = False
        if len(column_types) > 1 and column_types[0] == TYPE_NAME_TO_CODE["Int"]:
            if all(ct in [TYPE_NAME_TO_CODE["String"], TYPE_NAME_TO_CODE["StringID"]] for ct in column_types[1:]):
                use_cc_padding_after_first_int = True
                print("INFO: Detectado padrão [Int, String, ...]. O padding após o primeiro Int será preenchido com 0xCC.")

        # --- MUDANÇA AQUI: Trata Int e IntID da mesma forma no cálculo ---
        unaligned_row_size = 0
        for col_type in column_types:
            size = 4 if col_type in [TYPE_NAME_TO_CODE["Int"], TYPE_NAME_TO_CODE["IntID"]] else 8
            alignment = size
            padding = (alignment - (unaligned_row_size % alignment)) % alignment
            unaligned_row_size += padding + size
        row_size = unaligned_row_size
        if row_size % 8 != 0:
            row_size += 8 - (row_size % 8)

        tab_name_padded, tab_name_size = get_padded_string(ws.title)
        header = [b'EXPA', struct.pack('<II', 1, tab_name_size), tab_name_padded, struct.pack('<I', len(column_types))]
        header.extend([struct.pack('<I', ct) for ct in column_types])
        header.extend([struct.pack('<II', row_size, len(all_rows_data))])
        header_blob = b"".join(header)
        
        header_padding_needed = (8 - (len(header_blob) % 8)) % 8
        header_padding_bytes = b'\x00' * header_padding_needed
        data_start_offset = len(header_blob) + header_padding_needed

        data_blob_list, chnk_strings = [], []
        abs_offset = data_start_offset
        for row_data in all_rows_data:
            row_bytes, offset_in_row = [], 0
            for col_idx, (val, col_type) in enumerate(zip(row_data, column_types)):
                # --- MUDANÇA AQUI: Trata Int e IntID da mesma forma na escrita ---
                size = 4 if col_type in [TYPE_NAME_TO_CODE["Int"], TYPE_NAME_TO_CODE["IntID"]] else 8
                alignment = size
                padding = (alignment - (offset_in_row % alignment)) % alignment

                if padding > 0:
                    if use_cc_padding_after_first_int and col_idx == 1:
                        row_bytes.append(b'\xCC' * padding)
                    else:
                        row_bytes.append(b'\x00' * padding)
                
                if col_type in [TYPE_NAME_TO_CODE["Int"], TYPE_NAME_TO_CODE["IntID"]]:
                    row_bytes.append(struct.pack('<i', val))
                else: # String ou StringID
                    target_offset = abs_offset + offset_in_row + padding
                    if val: chnk_strings.append((target_offset, val))
                    row_bytes.append(b'\x00' * 8)
                
                offset_in_row += padding + size
            
            final_padding = row_size - offset_in_row
            if final_padding > 0:
                row_bytes.append(b'\x00' * final_padding)

            data_blob_list.append(b"".join(row_bytes))
            abs_offset += row_size
            
        data_blob = b"".join(data_blob_list)

        chnk_parts = [b'CHNK', struct.pack('<I', len(chnk_strings))]
        for offset, s_val in chnk_strings:
            padded_s, padded_size = get_padded_string(s_val)
            chnk_parts.extend([struct.pack('<II', offset, padded_size), padded_s])
        
        with open(mbe_path, 'wb') as f:
            f.write(header_blob)
            f.write(header_padding_bytes)
            f.write(data_blob)
            f.write(b"".join(chnk_parts))
        print("\nRecompilação concluída com sucesso!")
    except Exception as e:
        print(f"Ocorreu um erro durante a recompilação: {e}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Recompila .XLSX para .MBE.")
    parser.add_argument("input_file"); parser.add_argument("-o", "--output")
    args = parser.parse_args()
    recompile_mbe(args.input_file, args.output or f"{os.path.splitext(os.path.basename(args.input_file))[0]}.mbe")